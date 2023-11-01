import openpyxl

# Replace these with your input Excel file, sheet name, column letter, and constant value
input_excel_file = "add.xlsx"
sheet_name = "Sheet1"
column_letter = "A"
constant = 15

# Define a new file name for the output Excel file
output_excel_file = "output_file.xlsx"

# Load the input Excel file
workbook = openpyxl.load_workbook(input_excel_file)
sheet = workbook[sheet_name]

# Create a new workbook for the output file
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Copy the original sheet to the new workbook
for row in sheet.iter_rows():
    output_sheet.append([cell.value for cell in row])

# Iterate through the rows in the specified column and add the constant value
for row in output_sheet.iter_rows(min_col=sheet[column_letter][0].column, max_col=sheet[column_letter][0].column):
    for cell in row:
        if cell.row != 1:  # Skip the header row if present
            cell.value += constant

# Save the modified data to the output Excel file
output_workbook.save(output_excel_file)

print(f"Added {constant} to column '{column_letter}' and saved to '{output_excel_file}'.")
